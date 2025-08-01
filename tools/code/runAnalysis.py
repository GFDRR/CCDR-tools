# Required libraries
import os, gc
import warnings
import numpy as np
import pandas as pd
import geopandas as gpd
import folium
from branca.colormap import LinearColormap
from openpyxl import load_workbook
import rasterio
import rioxarray as rxr
from rasterstats import gen_zonal_stats, zonal_stats

# Importing internal libraries
import common
import input_utils
from damageFunctions import FL_mortality_factor, FL_damage_factor_builtup, FL_damage_factor_agri, TC_damage_factor_builtup 

# Importing the libraries for parallel processing
import itertools as it
from functools import partial
import multiprocess as mp

DATA_DIR = common.DATA_DIR
OUTPUT_DIR = common.OUTPUT_DIR
warnings.filterwarnings("ignore", message="'GeoSeries.swapaxes' is deprecated", category=FutureWarning)

# Defining functions for parallel processing of zonal_stats
def chunks(iterable_data, n):
    it_data = iter(iterable_data)
    for chunk in iter(lambda: list(it.islice(it_data, n)), []):
        yield chunk

def zonal_stats_partial(feats, raster, stats="*", affine=None, nodata=None, all_touched=True):
    # Partial zonal stats for parallel processing on a list of features
    return zonal_stats(feats, raster, stats=stats, affine=affine, nodata=nodata, all_touched=all_touched)

def zonal_stats_parallel(args):
    # Zonal stats for a parallel processing on a list of features
    return zonal_stats_partial(*args)

def save_excel_file(excel_file, dataset, sheet_name, summary_sheet=False):

    if os.path.exists(excel_file):
        excel_writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace')
        summary_exists = 'Summary' in load_workbook(excel_file, read_only=True).sheetnames
    else:
        excel_writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        
    if summary_sheet and summary_exists:
        summary_df = pd.read_excel(excel_file, sheet_name='Summary')
        dataset = merge_dfs(summary_df, dataset)
        
    with excel_writer:
        dataset.to_excel(excel_writer, sheet_name=sheet_name, index=False)
        

def merge_dfs(df_left, df_right, on_columns=['RP', 'Freq', 'Ex_freq'], how='outer'):
    
    # Get the unique columns from both DataFrames
    all_columns = list(df_left.columns.union(df_right.columns, sort=False))

    # Perform the merge
    merged_df = pd.merge(df_left, df_right, on=on_columns, how=how, suffixes=('_x', '_y'))
    common_columns = [col for col in df_left.columns if col in df_right.columns and col not in on_columns]

    for col in common_columns:
        merged_df[col] = merged_df[f'{col}_x'].combine_first(merged_df[f'{col}_y'])
        merged_df.drop([f'{col}_x', f'{col}_y'], axis=1, inplace=True)
        
    merged_df = merged_df.groupby(on_columns, as_index=False).first()
        
    return merged_df[all_columns]


# Process exposure data
def process_exposure_data(country, haz_type, exp_cat, exp_nam, exp_year, exp_folder):

    try:
        if exp_nam is not None:
            # Use the custom exposure data if provided
            exp_ras = f"{exp_folder}/{exp_nam}.tif"
            if not os.path.exists(exp_ras):
                raise FileNotFoundError(f"Custom exposure data not found: {exp_ras}")
                 
        else:
            # Use default exposure data based on exp_cat
            if exp_cat == 'POP':
                exp_ras = f"{exp_folder}/{country}_POP.tif"
                if not os.path.exists(exp_ras):
                    print(f"Population data not found. Fetching data for {country}...")
                    input_utils.fetch_population_data(country, exp_year)
                    if not os.path.exists(exp_ras):
                        raise FileNotFoundError(f"Failed to fetch population data for {country}")
            elif exp_cat == 'BU':
                exp_ras = f"{exp_folder}/{country}_BU.tif"
                if not os.path.exists(exp_ras):
                    print(f"Built-up data not found. Fetching data for {country}...")
                    input_utils.fetch_built_up_data(country)
                    if not os.path.exists(exp_ras):
                        raise FileNotFoundError(f"Failed to fetch built-up data for {country}")
            elif exp_cat == 'AGR':
                exp_ras = f"{exp_folder}/{country}_AGR.tif"

                if not os.path.exists(exp_ras):
                    print(f"Agriculture data not found. Fetching data for {country}...")
                    input_utils.fetch_agri_data(country)
                    if not os.path.exists(exp_ras):
                        raise FileNotFoundError(f"Failed to fetch agricultural data for {country}")
            else:
                raise ValueError(f"Missing or unknown exposure category: {exp_cat}")

        if not os.path.exists(exp_ras):
            raise FileNotFoundError(f"Exposure raster not found after processing: {exp_ras}")

        # Assign a default damage factor based on the haz_type and exp_cat
        if haz_type == 'FL':
            if exp_cat == 'POP':
                damage_factor = FL_mortality_factor
            elif exp_cat == 'BU':
                damage_factor = FL_damage_factor_builtup
            elif exp_cat == 'AGR':
                damage_factor = FL_damage_factor_agri
        elif haz_type == 'TC':
            if exp_cat == 'BU':
                damage_factor = TC_damage_factor_builtup
            else:
                damage_factor = lambda x, _: x
        else:
                raise ValueError(f"Unknown hazard type: {haz_type}")

        return exp_ras, damage_factor

    except Exception as e:
        print(f"Error in process_exposure_data: {str(e)}")


def exception_handler(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            print(f"An error occurred: {e}")
            return None
            
    return wrapper


# Defining the main function to run the analysis
@exception_handler
def run_analysis(
    country: str, haz_type: str, haz_cat: str, period: str, scenario: str,
    valid_RPs: list[int], min_haz_threshold: float, exp_cat: str,
    exp_nam: str, exp_year: str, adm_level: str, analysis_type: str, 
    class_edges: list[float], save_check_raster: bool, n_cores: int = None,
    use_custom_boundaries=False, custom_boundaries_file_path=None, custom_code_field=None,
    custom_name_field=None, wb_region=None
):
    """
    Run specified analysis.

    Parameters
    ----------
    country : country ISOa3 code
    haz_type : hazard type 'FL' for floods or 'TC' for tropical cyclones
    haz_cat : hazard category
    period: time period
    scenario: SSP scenario
    valid_RPs : return period values to be considered
    min_haz_threshold : minimum value for hazard values
    exp_cat : exposure category
    exp_nam : exposure user-specified file name or source
    exp_year: exposure year of reference
    adm_level : ADM level of sub-national boundaries
    analysis_type : type of analysis (class or function)
    class_edges : class edges for class-based analysis
    save_check_raster : save intermediate results to disk?
    """

    try:
        # Defining the location of administrative, hazard and exposure folders
        if haz_type == 'TC':  # Strong Wind (Tropical Cyclones)
            haz_folder = f"{DATA_DIR}/HZD/GLB/STORM/{period}"
        elif haz_type == 'FL':  # Floods (FLUVIAL_UNDEFENDED, FLUVIAL_DEFENDED, etc.)
            haz_folder = f"{DATA_DIR}/HZD/{country}/{haz_cat}/{period}/{scenario.replace('-', '_')}"

        exp_folder = f"{DATA_DIR}/EXP"

        # Validating Classes analysis parameters
        if analysis_type == "Classes":
            if not class_edges:
                raise ValueError("Class edges must be provided for Classes analysis")
            is_seq = np.all(np.diff(class_edges) > 0)
            if not is_seq:
                raise ValueError("Class thresholds are not sequential. Lower classes must be less than class thresholds above.")
            bin_seq = class_edges + [np.inf]
            num_bins = len(bin_seq)
        else:
            bin_seq = None
            num_bins = None

        # Fetch the ADM data
        if use_custom_boundaries:
            print(f"Using custom boundaries from file: {custom_boundaries_file_path}")
            adm_data = gpd.read_file(custom_boundaries_file_path)
            code_field = custom_code_field
            name_field = custom_name_field
            all_adm_codes = [code_field]
            all_adm_names = [name_field]
        else:
            print(f"Fetching ADM data for {country}, level {adm_level}")
            adm_data = input_utils.get_adm_data(country, adm_level)
            field_names = common.adm_field_mapping.get(adm_level, {})
            code_field = field_names.get('code')
            name_field = field_names.get('name')
            all_adm_codes = adm_data.columns[adm_data.columns.str.contains(r"HASC_\d$")].to_list()
            all_adm_names = adm_data.columns[adm_data.columns.str.contains(r"NAM_\d$")].to_list()

        if not (code_field and name_field):
            raise ValueError(f"Field names for ADM level {adm_level} not found")

        # Handle exposure data
        print(f"Processing exposure data for {exp_cat}")
        exp_ras, damage_factor = process_exposure_data(country, haz_type, exp_cat, exp_nam, exp_year, exp_folder)

        # Importing the exposure data
        # Open the raster dataset
        with rasterio.open(exp_ras) as src:
            original_nodata = src.nodata
        exp_data = rxr.open_rasterio(exp_ras)[0].astype('float32')  # Open exposure dataset
        # Handle nodata values
        if original_nodata is not None:
            # Mask the original nodata values
            exp_data = exp_data.where(exp_data != original_nodata)
        exp_data.rio.write_nodata(-1.0, inplace=True)
        exp_data.data[exp_data < 0.0] = 0.0

        # Parallel processing setup
        cores = min(len(valid_RPs), mp.cpu_count()) if n_cores is None else n_cores

        with mp.Pool(cores) as p:
            # Get total exposure for each ADM area
            func = partial(zonal_stats_partial, raster=exp_ras, stats="sum")
            stats_parallel = p.map(func, np.array_split(adm_data.geometry, cores))

        exp_per_ADM = list(it.chain(*stats_parallel))

        # Creating the results pandas dataframe
        result_df = adm_data.loc[:, all_adm_codes + all_adm_names + ["geometry"]]
        result_df[f"ADM{adm_level}_{exp_cat}"] = [x['sum'] for x in exp_per_ADM]

        # Cleaning-up memory
        del (stats_parallel, exp_per_ADM)
        gc.collect()
        
        # Defining the list of valid prob_RPs - probability of return period
        prob_RPs = 1./np.array(valid_RPs)
        prob_RPs_LB = np.append(-np.diff(prob_RPs), prob_RPs[-1]).tolist()           # Lower bound - alternative --> prob_RPs_LB = np.append(1-prob_RPs[0],-np.diff(prob_RPs)).tolist() # Lower bound [0-1]
        prob_RPs_UB = np.insert(-np.diff(prob_RPs), 0, 0.).tolist()                  # Upper bound - alternative --> prob_RPs_UB = np.append(1-prob_RPs[0],-np.diff(prob_RPs)).tolist() # Upper bound [0-1]
        prob_RPs_Mean = ((np.array(prob_RPs_LB) + np.array(prob_RPs_UB))/2).tolist() # Mean value
        prob_RPs_df = pd.DataFrame({'RPs':valid_RPs,
                                    'prob_RPs':prob_RPs,
                                    'prob_RPs_LB':prob_RPs_LB,
                                    'prob_RPs_UB':prob_RPs_UB,
                                    'prob_RPs_Mean':prob_RPs_Mean})
        prob_RPs_df.to_csv(os.path.join(OUTPUT_DIR, f"{country}_{haz_cat}_prob_RPs.csv"), index=False)
        
        # Computing the results for each RP
        n_valid_RPs_gt_1 = len(valid_RPs) > 1
        cores = min(len(valid_RPs), mp.cpu_count()) if n_cores is None else n_cores
        with mp.Pool(cores) as p:
            params = {
                "haz_folder": haz_folder,
                "analysis_type": analysis_type,
                "country": country,
                "haz_cat": haz_cat,
                "period": period,
                "scenario": scenario,
                "exp_cat": exp_cat,
                "exp_data": exp_data,
                "min_haz_threshold": min_haz_threshold,
                "damage_factor": damage_factor,
                "save_check_raster": save_check_raster,
                "bin_seq": bin_seq,
                "num_bins": num_bins,
                "adm_data": adm_data,
            }
            func = partial(calc_imp_RPs, wb_region=wb_region, **params)
            res = p.map(func, np.array_split(valid_RPs, cores))
        if not isinstance(res, list):
            to_concat = [result_df, res]
        else:
            to_concat = [result_df] + res

        result_df = pd.concat(to_concat, axis=1) # Concatenating the results
        result_df = result_df.replace(np.nan, 0) # Converting eventual nan/null to zero
        result_df = result_df_reorder_columns(result_df, valid_RPs, analysis_type, exp_cat, 
                                            adm_level, all_adm_codes, all_adm_names)
        result_df = calc_EAEI(result_df, valid_RPs, prob_RPs_df, 'LB', 
                            analysis_type, exp_cat, adm_level, num_bins, n_valid_RPs_gt_1)
        result_df = calc_EAEI(result_df, valid_RPs, prob_RPs_df, 'UB', 
                            analysis_type, exp_cat, adm_level, num_bins, n_valid_RPs_gt_1)
        result_df = calc_EAEI(result_df, valid_RPs, prob_RPs_df, 'Mean', 
                            analysis_type, exp_cat, adm_level, num_bins, n_valid_RPs_gt_1)
        result_df = result_df.round(3) # Round to three decimal places to avoid giving the impression of high precision
        
        # If method == 'Mean', then simplify it's name    
        # If not n_valid_RPs_gt_1 and any column contains the initial part as 'RP1_', it is removed then
        replace_string = '_Mean' if n_valid_RPs_gt_1 else 'RP1'
        result_df_colnames = [s.replace(replace_string, '') for s in result_df.columns]
        result_df.columns = result_df_colnames
        
        # Write output csv table and geopackages
        save_geopackage(result_df, country, adm_level, haz_cat, exp_cat, period, scenario, analysis_type, valid_RPs)
        
        # Trying to fix output not being passed to plot_results
        return result_df

    except Exception as e:
        print(f"An error occurred in run_analysis: {str(e)}")
        raise    

def calc_imp_RPs(RPs, haz_folder, analysis_type, country, haz_cat, period, scenario, exp_cat, exp_data, min_haz_threshold,
                 damage_factor, save_check_raster, bin_seq, num_bins, adm_data, wb_region):
    """
    Apply calculates for each given return period.
    """
    result_df = pd.DataFrame()
    for rp in RPs:
        if rp%1==0: rp = int(rp)
        # Loading the corresponding hazard dataset
        try:
            # We reproject using WarpedVRT as this applies the operation from disk
            # https://github.com/corteva/rioxarray/discussions/207
            # https://rasterio.readthedocs.io/en/latest/api/rasterio.vrt.html
            with rasterio.open(os.path.join(haz_folder, f"1in{rp}.tif")) as src:
                vrt_options = {
                    'src_crs': src.crs,
                    'crs': exp_data.rio.crs,
                    'transform': exp_data.rio.transform(recalc=True),
                    'height': exp_data.rio.height,
                    'width': exp_data.rio.width,
                }
                with rasterio.vrt.WarpedVRT(src, **vrt_options) as vrt:
                    haz_data = rxr.open_rasterio(vrt)[0]
                    haz_data.rio.write_nodata(-1.0, inplace=True)

        except rasterio._err.CPLE_OpenFailedError:
            raise IOError(f"Error occurred trying to open raster file: 1in{rp}.tif")

        # Set values below min threshold to nan
        haz_data = haz_data.where(haz_data.data > min_haz_threshold, np.nan)
        
        # Checking the analysis_type
        if analysis_type == "Function":
            # Assign impact factor (this is F_i in the equations)
            haz_data = damage_factor(haz_data, wb_region)
            if save_check_raster:
                haz_data.rio.to_raster(os.path.join(OUTPUT_DIR, f"{country}_{haz_cat}_{period}_{scenario}_{rp}_{exp_cat}_haz_imp_factor.tif"))
        elif analysis_type == "Classes":
            # Assign bin values to raster data - Follows: x_{i-1} <= x_{i} < x_{i+1}
            bin_idx = np.digitize(haz_data, bin_seq)

        # Calculate affected exposure in ADM
        # Filter down to valid areas affected areas which have people
        affected_exp = exp_data.where(haz_data.data > 0, np.nan)

        if save_check_raster:
            affected_exp.rio.to_raster(os.path.join(OUTPUT_DIR, f"{country}_{haz_cat}_{period}_{scenario}_{rp}_{exp_cat}_affected.tif"))
        
        # Conduct analyses for classes
        if analysis_type == "Classes":
            del haz_data
            for bin_x in reversed(range(num_bins)):
                # Compute the impact for this class
                impact_class = gen_zonal_stats(vectors=adm_data["geometry"],
                                               raster=np.array(bin_idx == bin_x).astype(int) * affected_exp.data,
                                               stats=["sum"], affine=affected_exp.rio.transform(), nodata=np.nan)
                result_df[f"RP{rp}_{exp_cat}_C{bin_x}_exp"] = [x['sum'] for x in impact_class]
                # Compute the cumulative impact for this class
                if bin_x < (num_bins - 1):
                    result_df[f"RP{rp}_{exp_cat}_C{bin_x}_exp"] = result_df[f"RP{rp}_{exp_cat}_C{bin_x}_exp"] + \
                                                                  result_df[f"RP{rp}_{exp_cat}_C{bin_x+1}_exp"]

        # Conduct analyses for function
        if analysis_type == "Function":
            # Compute the exposure per ADM level
            affected_exp_per_ADM = gen_zonal_stats(vectors=adm_data["geometry"], raster=affected_exp.data,
                                                   stats=["sum"], affine=affected_exp.rio.transform(), nodata=np.nan)
            result_df[f"RP{rp}_{exp_cat}_exp"] = [x['sum'] for x in affected_exp_per_ADM]
            # Calculate impacted exposure in affected areas
            impact_exp = affected_exp.data * haz_data
            # If save intermediate to disk is TRUE, then
            if save_check_raster:
                impact_exp.rio.to_raster(os.path.join(OUTPUT_DIR, f"{country}_{period}_{scenario}_{rp}_{exp_cat}_impact.tif"))
            # Compute the impact per ADM level
            impact_exp_per_ADM = gen_zonal_stats(vectors=adm_data["geometry"], raster=impact_exp.data, stats=["sum"],
                                                 affine=impact_exp.rio.transform(), nodata=np.nan)
            result_df[f"RP{rp}_{exp_cat}_imp"] = [x['sum'] for x in impact_exp_per_ADM]
            del (haz_data, impact_exp, impact_exp_per_ADM, affected_exp_per_ADM)

        del affected_exp
        gc.collect()

    return result_df

def result_df_reorder_columns(result_df, RPs, analysis_type, exp_cat, adm_level, all_adm_codes, all_adm_names):
    """
    Reorders the columns of result_df.
    """
    # Re-ordering and dropping selected columns for better presentation of the results
    
    if analysis_type != "Function":
        return result_df
    
    adm_column = f"ADM{adm_level}_{exp_cat}"
    
    all_RPs = ["RP" + str(rp) for rp in RPs]
    all_exp = [x + f"_{exp_cat}_exp" for x in all_RPs]
    all_imp = [x + f"_{exp_cat}_imp" for x in all_RPs]
    col_order = all_adm_codes + all_adm_names + [adm_column] + all_exp + all_imp + ["geometry"]
    result_df = result_df.loc[:, col_order]

    return result_df

def calc_EAEI(result_df, RPs, prob_RPs_df, method, analysis_type, exp_cat, 
              adm_level, num_bins, n_valid_RPs_gt_1):
    """
    Computes the EAE/EAI over each given return period.
    """
    for rp in RPs:

        # Exceedance Probability of return period
        freq = float(prob_RPs_df.loc[prob_RPs_df['RPs'] == rp, f'prob_RPs_{method}'].iloc[0])

        # Conduct analyses for classes
        if analysis_type == "Classes":
            # Compute the EAE for classes, if probabilistic (len(valid_RPs)>1)
            if n_valid_RPs_gt_1:
                for bin_x in reversed(range(num_bins)):
                    result_df[f"RP{rp}_{exp_cat}_C{bin_x}_EAE_tmp"] = result_df[f"RP{rp}_{exp_cat}_C{bin_x}_exp"] * freq

        # Conduct analyses for function
        if analysis_type == "Function":
            # Compute the EAI for this RP, if probabilistic (len(valid_RPs)>1)
            if n_valid_RPs_gt_1:
                result_df[f"RP{rp}_EAI_tmp"] = result_df[f"RP{rp}_{exp_cat}_imp"] * freq
        
    # Computing the EAE or EAI, if probabilistic (len(valid_RPs)>1)
    if n_valid_RPs_gt_1:
        # Computing EAE if analysis is Classes
        if analysis_type == "Classes":
            # Sum all EAI to get total EAI across all RPs and Classes
            for bin_x in reversed(range(0, num_bins)):
                C_EAE_cols = result_df.columns.str.contains(f"{exp_cat}_C{bin_x}_EAE_tmp")
                result_df.loc[:, f"{exp_cat}_C{bin_x}_EAE_{method}"] = result_df.loc[:, C_EAE_cols].sum(axis=1)

            # Calculate EAE% (Percent affected exposure per year)
            for bin_x in reversed(range(0, num_bins)):
                result_df.loc[:, f"{exp_cat}_C{bin_x}_EAE%_{method}"] = (result_df.loc[:,f"{exp_cat}_C{bin_x}_EAE_{method}"] / 
                                                                         result_df.loc[:,f"ADM{adm_level}_{exp_cat}"]) * 100.0
        # Computing EAI if analysis is Function
        if analysis_type == "Function":
            if n_valid_RPs_gt_1:
                # Sum all EAI to get total EAI across all RPs
                result_df.loc[:, f"{exp_cat}_EAI_{method}"] = result_df.loc[:,result_df.columns.str.contains('_EAI_tmp')].sum(axis=1)
                # Calculate Exp_EAI% (Percent affected exposure per year)
                result_df.loc[:, f"{exp_cat}_EAI%_{method}"] = (result_df.loc[:, f"{exp_cat}_EAI_{method}"] / 
                                                                result_df.loc[:,f"ADM{adm_level}_{exp_cat}"]) * 100.0

    # Dropping selected columns for better presentation of the results
    if analysis_type == "Function":
        all_EAI = [col for col in result_df.columns if '_EAI_tmp' in col] if n_valid_RPs_gt_1 else []
        result_df = result_df.drop(all_EAI, axis=1) # dropping
    if analysis_type == "Classes":
        all_EAE = [col for col in result_df.columns if '_EAE_tmp' in col] if n_valid_RPs_gt_1 else []
        result_df = result_df.drop(all_EAE, axis=1) # dropping
        
    return result_df

def create_summary_df(result_df, valid_RPs, exp_cat):
    summary_data = []
    for rp in valid_RPs:
        row = {'RP': rp, 'Freq': 1/rp}
        
        # Check for impact column
        impact_col = next((col for col in result_df.columns if f'RP{rp}_{exp_cat}_imp' in col), None)
        if impact_col:
            row[f'{exp_cat}_impact'] = result_df[impact_col].sum()
        
        summary_data.append(row)
    
    summary_df = pd.DataFrame(summary_data)
    
    # Calculate Ex_freq
    summary_df['Ex_freq'] = summary_df['Freq'].diff().abs().shift(-1)
    summary_df.loc[summary_df.index[-1], 'Ex_freq'] = summary_df.loc[summary_df.index[-1], 'Freq']
    
    # Calculate EAI
    if f'{exp_cat}_impact' in summary_df.columns:
        summary_df[f'{exp_cat}_EAI'] = summary_df[f'{exp_cat}_impact'] * summary_df['Ex_freq']
    
    return summary_df

def save_geopackage(result_df, country, adm_level, haz_cat, exp_cat, period, scenario, analysis_type, valid_RPs):
    # Ensure that the geometry column is correctly recognized
    if 'geometry' not in result_df.columns and 'geom' in result_df.columns:
        result_df = result_df.rename(columns={'geom': 'geometry'})
    elif 'geometry' not in result_df.columns:
        raise ValueError("The DataFrame does not contain a geometry column.")
   
    # Convert to GeoDataFrame if it's not already one
    if not isinstance(result_df, gpd.GeoDataFrame):
        result_df = gpd.GeoDataFrame(result_df, geometry='geometry')
   
    # Set the CRS to EPSG:4326
    result_df.set_crs(epsg=4326, inplace=True)
   
    # Remove the geometry column for the Excel export
    df_cols = result_df.columns
    no_geom = result_df.loc[:, df_cols[~df_cols.isin(['geometry'])]].fillna(0)
   
    # Prepare Excel writer
    if period == '2020':
        file_prefix = f"{country}_ADM{adm_level}_{haz_cat}_{period}"   
    else:
        file_prefix = f"{country}_ADM{adm_level}_{haz_cat}_{period}_{scenario}"

    # Create Excel writer object
    excel_file = os.path.join(common.OUTPUT_DIR, f"{file_prefix}.xlsx")
    
    if analysis_type == "Function":
        EAI_string = "EAI_" if len(valid_RPs) > 1 else ""
        sheet_name = f"{exp_cat}_{EAI_string}function"
        save_excel_file(excel_file, no_geom, sheet_name)
    elif analysis_type == "Classes":
        EAE_string = "EAE_" if len(valid_RPs) > 1 else ""
        sheet_name = f"{exp_cat}_{EAE_string}class"
        save_excel_file(excel_file, no_geom, sheet_name)
    else:
        raise ValueError("Unknown analysis type. Use 'Function' or 'Classes'.")
        
    return result_df  # Return the GeoDataFrame

def plot_results(result_df, exp_cat, analysis_type):
    # Convert result_df to GeoDataFrame if it's not already
    if not isinstance(result_df, gpd.GeoDataFrame):
        result_df = gpd.GeoDataFrame(result_df, geometry='geometry')
    
    # Determine the column to plot based on analysis_type
    if analysis_type == "Function":
        column = f'{exp_cat}_EAI'
    elif analysis_type == "Classes":
        column = f'RP10_{exp_cat}_C1'
    else:
        print("Unknown analysis approach")
        return None, None

    # Ensure the CRS is EPSG:4326
    result_df = result_df.to_crs(epsg=4326)

    # Filter out zero and negative values for color scaling
    non_zero_data = result_df[result_df[column] > 0]
    
    if len(non_zero_data) == 0:
        return None, None

    vmin = non_zero_data[column].min()
    vmax = non_zero_data[column].max()
    
    # Create a custom colormap
    colors = ['#FFEDA0', '#FED976', '#FEB24C', '#FD8D3C', '#FC4E2A', '#E31A1C', '#BD0026', '#800026']
    colormap = LinearColormap(colors=colors, vmin=vmin, vmax=vmax)
    
    # Create a style function that colors features based on their value
    def style_function(feature):
        value = feature['properties'][column]
        if value <= 0:
            return {
                'fillColor': 'transparent',
                'fillOpacity': 0,
                'color': 'black',
                'weight': 1,
            }
        return {
            'fillColor': colormap(value),
            'fillOpacity': 0.7,
            'color': 'black',
            'weight': 1,
        }
    
    # Create the GeoJson layer
    geojson_layer = folium.GeoJson(
        result_df,
        style_function=style_function,
        name=f"{exp_cat} - {column}"
    )
    
    return geojson_layer, colormap


def prepare_excel_gpkg_files(country, adm_level, haz_cat, period, scenario):

    file_prefix = f"{country}_ADM{adm_level}_{haz_cat}_{period}"

    if period != '2020':
        file_prefix += f"_{scenario}"

    excel_file = os.path.join(common.OUTPUT_DIR, f"{file_prefix}.xlsx")
    gpkg_file = os.path.join(common.OUTPUT_DIR, f"{file_prefix}.gpkg")

    return excel_file, gpkg_file


def prepare_sheet_name(analysis_type, return_periods, exp_cat):
    
    # Save results to Excel and GeoPackage
    if analysis_type == "Function":
        EAI_string = "EAI_" if len(return_periods) > 1 else ""
        sheet_name = f"{exp_cat}_{EAI_string}function"
    elif analysis_type == "Classes":
        EAE_string = "EAE_" if len(return_periods) > 1 else ""
        sheet_name = f"{exp_cat}_{EAE_string}class"
    else:
        raise ValueError("Unknown analysis type. Use 'Function' or 'Classes'.")    
    return sheet_name


def saving_excel_and_gpgk_file(result_df, excel_file, sheet_name, gpkg_file, exp_cat):
    # Save to Excel
    df_to_save = result_df.drop('geometry', axis=1, errors='ignore')
    save_excel_file(excel_file, df_to_save, sheet_name, summary_sheet=False)
                
    # Save to GeoPackage
    if isinstance(result_df, gpd.GeoDataFrame):
        result_df.to_file(gpkg_file, layer=sheet_name, driver='GPKG')
    else:
        print(f"Warning: Result for {exp_cat} is not a GeoDataFrame. Skipping GeoPackage export for this layer.")


def prepare_and_save_summary_df(summary_dfs, exp_cat_list, excel_file, return_file:bool = False):
    combined_summary = summary_dfs[0].copy().round(3)
    for df in summary_dfs[1:]:
        combined_summary = pd.merge(combined_summary, df.round(3), on=['RP', 'Freq', 'Ex_freq'], how='outer')

    # Reorder columns
    ordered_columns = ['RP', 'Freq', 'Ex_freq']
    for exp_cat in exp_cat_list:
        ordered_columns.extend([f'{exp_cat}_impact', f'{exp_cat}_EAI'])

    # Ensure all expected columns are present, fill with NaN if missing
    for col in ordered_columns:
        if col not in combined_summary.columns:
            combined_summary[col] = np.nan

    # Select only the ordered columns
    combined_summary = combined_summary[ordered_columns]

    # Save combined summary to Excel
    save_excel_file(excel_file, combined_summary, sheet_name='Summary', summary_sheet=True)
    
    if return_file:
        return combined_summary
